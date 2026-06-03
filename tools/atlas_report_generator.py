#!/usr/bin/env python3
"""Generate ATLAS-ready CSVs from monthly RISE source reports.

The generator scans Excel/PDF reports, detects supported source types, maps
community aliases to ATLAS Community_ID values, and writes one CSV per ATLAS
target table plus exception and manifest files.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple


try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - runtime dependency message
    raise SystemExit(
        "Missing dependency: openpyxl. Install it or run with the Codex bundled Python runtime."
    ) from exc

try:
    from pypdf import PdfReader
except ImportError:  # PDFs are optional; Excel parsing should still work.
    PdfReader = None


MONTHLY_FIELDS = [
    "Report_Month",
    "Community_ID",
    "Community_Name",
    "Region",
    "Market",
    "Asset_Type",
    "PMS_Property_Code",
    "Total_Units",
    "Occupied_Units",
    "Leased_Units",
    "Preleased_Units",
    "Vacant_Units",
    "Physical_Occupancy_%",
    "Leased_%",
    "Economic_Occupancy_%",
    "Avg_Market_Rent",
    "Actual_Rent",
    "Concessions_$",
    "Delinquency_$",
    "Gross_Potential_Rent_$",
    "Collected_Rent_$",
    "Other_Income_$",
    "Operating_Expenses_$",
    "NOI_$",
    "Traffic",
    "Leads",
    "Tours",
    "Applications",
    "Approved_Applications",
    "Move_Ins",
    "Move_Outs",
    "Renewals_Sent",
    "Renewals_Signed",
    "Reviews_Count",
    "Avg_Review_Rating",
    "Open_Work_Orders",
    "Completed_Work_Orders",
    "Period_Status",
    "Source_System",
    "Source_File_Name",
    "Agent_Notes",
    "Last_Updated",
    "Row_Status",
    "Error_Detail",
]

TABLE_SCHEMAS = {
    "monthly_upload": MONTHLY_FIELDS,
    "boxscore_floor_plan_monthly": [
        "Report_Month",
        "Community_ID",
        "Unit_Type",
        "Avg_SQFT",
        "Avg_Market_Rent",
        "Avg_Scheduled_Rent",
        "Avg_Net_Effective_Rent",
        "Units",
        "Excluded",
        "Rentable_Units",
        "Occupied_Units",
        "Vacant_Units",
        "Available_Units",
        "Occupied_No_Notice",
        "Notice_Rented",
        "Notice_Unrented",
        "Vacant_Rented",
        "Vacant_Unrented",
        "Occupancy_Pct",
        "Leased_Pct",
        "Trend",
        "Exposure_Pct",
        "Total_Square_Feet",
        "Total_Market_Rent",
        "Total_Scheduled_Charges",
        "Total_Budgeted_Rent",
        "Total_Scheduled_Rent",
        "Total_Scheduled_Other",
        "Total_Effective_Rent",
        "Total_Actual_Rent",
        "Vacant_Rented_Ready_Units",
        "Vacant_Rented_Not_Ready_Units",
        "Vacant_Unrented_Ready_Units",
        "Vacant_Unrented_Not_Ready_Units",
        "Leased_Units",
        "Exposure_Units",
        "Not_Exposed_Leased_Units",
        "Source_File_Name",
        "Source_Sheet",
    ],
    "boxscore_pulse_monthly": [
        "Report_Month",
        "Community_ID",
        "Unit_Type",
        "Units",
        "Move_Ins",
        "MTM",
        "MTM_Conversions",
        "Renewal_Leases_Approved",
        "Transfers",
        "Notices",
        "Move_Outs",
        "Renewal_Transfers",
        "Skips",
        "Evictions",
        "Leased",
        "Source_File_Name",
        "Source_Sheet",
    ],
    "boxscore_lead_monthly": [
        "Report_Month",
        "Community_ID",
        "Unit_Type",
        "New_Leads",
        "Email_Leads",
        "Call_Leads",
        "Online_Leads",
        "Walk_In_Leads",
        "Off_Site_Event_Leads",
        "Chat_Leads",
        "Text_Leads",
        "Other_Leads",
        "Emails",
        "Calls",
        "Chats",
        "Texts",
        "First_Visits_Tours",
        "Source_File_Name",
        "Source_Sheet",
    ],
    "boxscore_application_monthly": [
        "Report_Month",
        "Community_ID",
        "Unit_Type",
        "Applications_Completed",
        "Applications_Partially_Completed",
        "Applications_Completed_Cancelled",
        "Applications_Denied",
        "Applications_Approved",
        "Applications_Approved_Cancelled",
        "Lease_Conversions_Completed",
        "Lease_Conversions_Cancelled",
        "Lease_Conversions_Approved",
        "Source_File_Name",
        "Source_Sheet",
    ],
    "boxscore_make_ready_monthly": [
        "Report_Month",
        "Community_ID",
        "Status",
        "Vacant_Rented",
        "Vacant_Unrented",
        "Total_Vacant",
        "Pct_Of_Total_Vacant",
        "Source_File_Name",
        "Source_Sheet",
    ],
    "market_property_snapshot": [
        "Snapshot_Date",
        "Report_Month",
        "Subject_Community_ID",
        "Source_Property_Name",
        "Is_Subject_Property",
        "Distance",
        "Year_Built",
        "Total_Units",
        "Exposure_Pct",
        "Leased_Pct",
        "Preleased_Pct",
        "Applications_Last_7_Days",
        "Applications_Last_30_Days",
        "Concessions",
        "Concession_Pct",
        "Total_Available_Units",
        "Total_Leased_Units",
        "Vacant_Units",
        "Asking_Rent",
        "Asking_Rent_PSF",
        "Effective_Rent",
        "Effective_Rent_PSF",
        "Avg_SqFt",
        "Source_File_Name",
    ],
    "market_floor_plan_snapshot": [
        "Snapshot_Date",
        "Report_Month",
        "Subject_Community_ID",
        "Source_Property_Name",
        "Floor_Plan",
        "Bed_Count",
        "Bath_Count",
        "Sq_Ft",
        "Total_Units",
        "Unit_Mix",
        "Leased_Units_Last_7_Days",
        "Leased_Days_On_Market",
        "Leased_Asking_Rent",
        "Leased_Asking_Rent_PSF",
        "Available_Units",
        "Available_Days_On_Market",
        "Available_Asking_Rent",
        "Available_Asking_Rent_PSF",
        "Last_7_Days_Asking_Rent_Trend",
        "Source_File_Name",
    ],
    "market_unit_snapshot": [
        "Snapshot_Date",
        "Report_Month",
        "Subject_Community_ID",
        "Source_Property_Name",
        "Address",
        "Floor_Plan",
        "Unit_Number",
        "Bed_Count",
        "Bath_Count",
        "Sq_Ft",
        "Unit_Status",
        "Date_First_Available",
        "Leased_Date",
        "Days_On_Market",
        "Times_Listed",
        "Asking_Rent",
        "Asking_Rent_PSF",
        "Effective_Rent",
        "Effective_Rent_PSF",
        "Concession_Pct",
        "Concession_Details",
        "Source_File_Name",
    ],
    "rent_roll_unit_monthly": [
        "Report_Month",
        "Community_ID",
        "Source_Property_Name",
        "Bldg_Unit",
        "Unit_Type",
        "SqFt",
        "Unit_Status",
        "Resident_Name",
        "Move_In",
        "Lease_Start",
        "Lease_End",
        "Expected_Move_Out",
        "Market_Rent_Budgeted",
        "Actual_Charges",
        "In_Place_Rent_Scheduled",
        "Balance",
        "Deposit_Held",
        "Ready_Status",
        "PII_Load_Flag",
        "Source_File_Name",
        "Source_Page",
    ],
    "renewal_tracker_monthly": [
        "Report_Month",
        "Community_ID",
        "Source_Property_Name",
        "Renewal_Month_Tab",
        "Resident_Name",
        "Unit",
        "Unit_Type",
        "Expiration_Date",
        "Notice_90_Day",
        "Notice_60_Day",
        "Notice_30_Day",
        "Deposit_Held",
        "Current_Rate",
        "Recommended_Offer",
        "Investor_Override_Increase_Pct",
        "Investor_Override_Offer",
        "Offer_1_Conservative",
        "Offer_2_Balanced",
        "Offer_3_Aggressive",
        "Signed_Offer",
        "Renewal_Signed_Date",
        "Transfer_Date",
        "NTV_Received_Date",
        "Phone",
        "Notes",
        "Market_Rate",
        "Budget_Rate",
        "Occupancy_Status",
        "Rent_Growth_Offer_1",
        "Rent_Growth_Offer_2",
        "Rent_Growth_Offer_3",
        "Signed_Rent_Growth",
        "PII_Load_Flag",
        "Source_File_Name",
        "Source_Sheet",
    ],
    "trending_occupancy_monthly": [
        "Report_Month",
        "Community_ID",
        "Date_Range",
        "Beginning_Occupancy",
        "Move_Ins",
        "Move_Outs",
        "Ending_Occupancy",
        "Block_Index",
        "Mapping_Method",
        "Source_File_Name",
    ],
    "audit_move_in_out": [
        "Report_Month",
        "Community_ID",
        "Source_File_Name",
        "Source_Move_Ins",
        "Source_Move_Outs",
        "Uploaded_Move_Ins",
        "Uploaded_Move_Outs",
        "Final_Move_Ins",
        "Final_Move_Outs",
        "Audit_Status",
        "Correction_Note",
    ],
    "source_file_log": [
        "Source_File_Name",
        "Source_Path",
        "Detected_Source_Type",
        "Detected_Period",
        "Report_Month",
        "Snapshot_Date",
        "Community_Scope",
        "Rows_Extracted",
        "Status",
        "Notes",
    ],
    "exceptions": [
        "Severity",
        "Source_File_Name",
        "Source_Location",
        "Issue",
        "Recommended_Action",
    ],
    "manifest": ["Table", "CSV_File", "Rows"],
}

COMMUNITIES = {
    "SER": {
        "Community_Name": "Sereno",
        "Region": "Florida",
        "Market": "Jacksonville",
        "Asset_Type": "Multifamily",
        "Total_Units": 320,
        "PMS_Property_Code": "SERENO",
        "In_Current_Atlas_Master": True,
    },
    "GKP": {
        "Community_Name": "Glen Kernan Park",
        "Region": "Florida",
        "Market": "Jacksonville",
        "Asset_Type": "Lease-Up",
        "Total_Units": 308,
        "PMS_Property_Code": "GKP",
        "In_Current_Atlas_Master": True,
    },
    "NOC": {
        "Community_Name": "Nocatee",
        "Region": "Florida",
        "Market": "Jacksonville",
        "Asset_Type": "Lease-Up",
        "Total_Units": 178,
        "PMS_Property_Code": "NOCATEE",
        "In_Current_Atlas_Master": True,
    },
    "FLO": {
        "Community_Name": "Florence Villa",
        "Region": "Florida",
        "Market": "Central Florida",
        "Asset_Type": "Lease-Up",
        "Total_Units": 224,
        "PMS_Property_Code": "FLOVILLA",
        "In_Current_Atlas_Master": True,
    },
    "BAR": {
        "Community_Name": "Bartram",
        "Region": "Florida",
        "Market": "Jacksonville",
        "Asset_Type": "Multifamily",
        "Total_Units": 297,
        "PMS_Property_Code": "BARTRAM",
        "In_Current_Atlas_Master": True,
    },
    "CIT": {
        "Community_Name": "Citrus Ridge",
        "Region": "Florida",
        "Market": "Central Florida",
        "Asset_Type": "Lease-Up",
        "Total_Units": 222,
        "PMS_Property_Code": "CITRIDGE",
        "In_Current_Atlas_Master": True,
    },
    "VIE": {
        "Community_Name": "Viera",
        "Region": "Florida",
        "Market": "Space Coast",
        "Asset_Type": "Lease-Up",
        "Total_Units": 166,
        "PMS_Property_Code": "VIERA",
        "In_Current_Atlas_Master": True,
    },
    "BAY": {
        "Community_Name": "Baymeadows",
        "Region": "Florida",
        "Market": "Jacksonville",
        "Asset_Type": "Lease-Up",
        "Total_Units": 331,
        "PMS_Property_Code": "BAYMEADOWS",
        "In_Current_Atlas_Master": True,
    },
    "DOR": {
        "Community_Name": "Doro",
        "Region": "Florida",
        "Market": "Jacksonville",
        "Asset_Type": "Multifamily",
        "Total_Units": 247,
        "PMS_Property_Code": "DORO",
        "In_Current_Atlas_Master": True,
    },
    "STA": {
        "Community_Name": "St. Augustine",
        "Region": "Florida",
        "Market": "St. Augustine",
        "Asset_Type": "Lease-Up",
        "Total_Units": 272,
        "PMS_Property_Code": "STAUGUSTINE",
        "In_Current_Atlas_Master": False,
    },
}

COMMUNITY_ALIASES = {
    "SER": ["Sereno", "RISE Sereno", "Rise Sereno"],
    "GKP": [
        "Glen Kernan Park",
        "RISE at Glen Kernan Park",
        "Rise At Glen Kernan Park",
        "RISE_at_Glen_Kernan_Park",
    ],
    "NOC": ["Nocatee", "RISE at Nocatee", "Rise At Nocatee"],
    "FLO": ["Florence Villa", "RISE Florence Villa Townhomes", "Rise Florence Villa Townhomes"],
    "BAR": ["Bartram", "RISE Bartram Park", "Rise Bartram Park"],
    "CIT": ["Citrus Ridge", "RISE Citrus Ridge Townhomes", "Rise Citrus Ridge Townhomes"],
    "VIE": ["Viera", "RISE Viera", "Rise Viera"],
    "BAY": ["Baymeadows", "RISE Baymeadows", "Rise Baymeadows"],
    "DOR": ["Doro", "RISE Doro", "Rise Doro"],
    "STA": ["St. Augustine", "RISE St. Augustine", "Rise St. Augustine"],
}

SOURCE_PRIORITY = {
    "Market Survey": 40,
    "Rent Roll PDF": 50,
    "Renewal Tracker": 70,
    "Trending Occupancy": 80,
    "Box Score": 90,
}
TRENDING_OCCUPANCY_COMMUNITY_ORDER = ["GKP", "NOC", "BAR", "BAY", "CIT", "DOR", "FLO", "SER", "STA", "VIE"]
TODAY = dt.date.today().isoformat()


def normalize_name(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("&", "and")
    text = re.sub(r"\b(rise|at|the|apartments|townhomes)\b", " ", text.lower())
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


ALIAS_TO_COMMUNITY: Dict[str, str] = {}
for community_id, aliases in COMMUNITY_ALIASES.items():
    for alias in aliases + [COMMUNITIES[community_id]["Community_Name"]]:
        ALIAS_TO_COMMUNITY[normalize_name(alias)] = community_id


def to_number(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text in {"--", "-", "N/A"}:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace(",", "").replace("$", "").replace("%", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return -number if negative else number


def to_intish(value: Any) -> Optional[int]:
    number = to_number(value)
    if number is None:
        return None
    return int(round(number))


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


MARKET_SURVEY_PROPERTY_HEADERS = {
    "Property Name",
    "Property Snapshot",
}

MARKET_SURVEY_SECTION_LABELS = {
    "Property Rent Comparison",
    "Property Metric Rankings",
    "Asking Rent",
    "Asking Rent / Sq Ft",
    "Effective Rent",
    "Effective Rent / Sq Ft",
    "Sq Ft",
    "Concessions",
    "Concession %",
    "Total Avail. Units",
    "Exposure %",
    "Leased %",
    "Applications Last 7 Days",
    "Applications Last 30 Days",
    "Total Leased Units",
    "Vacant Units",
}


def is_market_survey_property_row(source_property: str, distance: Any) -> bool:
    label = clean_text(source_property)
    if not label or label.lower().startswith("average"):
        return False
    if label in MARKET_SURVEY_PROPERTY_HEADERS:
        return False
    if label in MARKET_SURVEY_SECTION_LABELS:
        return False
    if label == "Property":
        return False
    return isinstance(distance, (int, float, dt.datetime, dt.date)) or to_number(distance) is not None


def iso_date(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = clean_text(value)
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"):
        try:
            return dt.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def first_day_of_month(date_value: dt.date) -> str:
    return dt.date(date_value.year, date_value.month, 1).isoformat()


def parse_report_month(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""

    match = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", text)
    if match:
        parsed = dt.datetime.strptime(match.group(1), "%m/%d/%Y").date()
        return first_day_of_month(parsed)

    match = re.search(
        r"\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+(\d{4})\b",
        text,
        flags=re.I,
    )
    if match:
        parsed = dt.datetime.strptime(f"{match.group(1)[:3]} {match.group(2)}", "%b %Y").date()
        return first_day_of_month(parsed)

    match = re.search(r"\b(20\d{2})[-_](\d{2})[-_]\d{2}\b", text)
    if match:
        return f"{match.group(1)}-{match.group(2)}-01"

    match = re.search(r"\b(20\d{2})[-_](\d{2})\b", text)
    if match:
        return f"{match.group(1)}-{match.group(2)}-01"

    return ""


def parse_snapshot_date_from_filename(path: Path) -> str:
    match = re.match(r"(20\d{2})_(\d{2})_(\d{2})_", path.stem)
    if match:
        return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
    match = re.search(r"(20\d{2})[-_](\d{2})[-_](\d{2})", path.stem)
    if match:
        return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
    return ""


def report_month_from_cli(value: str) -> str:
    if not value:
        return ""
    match = re.fullmatch(r"(20\d{2})-(\d{2})", value)
    if match:
        return f"{match.group(1)}-{match.group(2)}-01"
    parsed = parse_report_month(value)
    if parsed:
        return parsed
    raise SystemExit("--report-month must look like YYYY-MM, YYYY-MM-DD, or a parseable month.")


def map_community(name: Any) -> Optional[str]:
    normalized = normalize_name(name)
    if normalized in ALIAS_TO_COMMUNITY:
        return ALIAS_TO_COMMUNITY[normalized]
    # A tiny amount of forgiving matching for filenames such as RISE_at_Glen_Kernan_Park.
    for alias, community_id in ALIAS_TO_COMMUNITY.items():
        if alias and (alias in normalized or normalized in alias):
            return community_id
    return None


def community_base_row(report_month: str, community_id: str) -> Dict[str, Any]:
    community = COMMUNITIES[community_id]
    return {
        "Report_Month": report_month,
        "Community_ID": community_id,
        "Community_Name": community["Community_Name"],
        "Region": community["Region"],
        "Market": community["Market"],
        "Asset_Type": community["Asset_Type"],
        "PMS_Property_Code": community["PMS_Property_Code"],
        "Total_Units": community["Total_Units"],
        "Period_Status": "Actual",
        "Source_System": "",
        "Source_File_Name": "",
        "Agent_Notes": "",
        "Last_Updated": TODAY,
        "Row_Status": "OK" if community["In_Current_Atlas_Master"] else "REVIEW",
        "Error_Detail": "" if community["In_Current_Atlas_Master"] else "Community is not in the current Atlas Community_Master.",
    }


@dataclass
class RunState:
    monthly_rows: Dict[Tuple[str, str], Dict[str, Any]] = field(default_factory=dict)
    monthly_priority: Dict[Tuple[str, str], int] = field(default_factory=dict)
    detail_rows: Dict[str, List[Dict[str, Any]]] = field(
        default_factory=lambda: {table: [] for table in TABLE_SCHEMAS if table not in {"monthly_upload", "source_file_log", "exceptions", "manifest"}}
    )
    source_file_log: List[Dict[str, Any]] = field(default_factory=list)
    exceptions: List[Dict[str, Any]] = field(default_factory=list)

    def add_exception(
        self,
        severity: str,
        source_file: Path,
        location: str,
        issue: str,
        action: str,
    ) -> None:
        self.exceptions.append(
            {
                "Severity": severity,
                "Source_File_Name": source_file.name,
                "Source_Location": location,
                "Issue": issue,
                "Recommended_Action": action,
            }
        )

    def add_file_log(
        self,
        source_file: Path,
        source_type: str,
        detected_period: str = "",
        report_month: str = "",
        snapshot_date: str = "",
        community_scope: str = "",
        rows_extracted: int = 0,
        status: str = "OK",
        notes: str = "",
    ) -> None:
        self.source_file_log.append(
            {
                "Source_File_Name": source_file.name,
                "Source_Path": str(source_file),
                "Detected_Source_Type": source_type,
                "Detected_Period": detected_period,
                "Report_Month": report_month,
                "Snapshot_Date": snapshot_date,
                "Community_Scope": community_scope,
                "Rows_Extracted": rows_extracted,
                "Status": status,
                "Notes": notes,
            }
        )

    def merge_monthly(
        self,
        report_month: str,
        community_id: str,
        updates: Dict[str, Any],
        source_type: str,
        source_file: Path,
        notes: Sequence[str] = (),
        review: bool = False,
    ) -> None:
        if not report_month or not community_id:
            return
        key = (report_month, community_id)
        priority = SOURCE_PRIORITY.get(source_type, 0)
        if key not in self.monthly_rows:
            self.monthly_rows[key] = community_base_row(report_month, community_id)
            self.monthly_priority[key] = priority

        row = self.monthly_rows[key]
        current_priority = self.monthly_priority[key]
        should_override = priority >= current_priority
        changed = False

        for field_name, value in updates.items():
            if field_name not in MONTHLY_FIELDS or value in (None, ""):
                continue
            if should_override or row.get(field_name) in (None, ""):
                if row.get(field_name) != value:
                    changed = True
                    row[field_name] = value

        if should_override:
            self.monthly_priority[key] = priority

        source_names = {item.strip() for item in str(row.get("Source_File_Name", "")).split(";") if item.strip()}
        source_names.add(source_file.name)
        row["Source_File_Name"] = "; ".join(sorted(source_names))
        systems = {item.strip() for item in str(row.get("Source_System", "")).split(";") if item.strip()}
        systems.add(source_type if source_type != "Box Score" else "Excel Export")
        row["Source_System"] = "; ".join(sorted(systems))

        if notes:
            existing = [item.strip() for item in str(row.get("Agent_Notes", "")).split(" | ") if item.strip()]
            for note in notes:
                if note and note not in existing:
                    existing.append(note)
            row["Agent_Notes"] = " | ".join(existing)

        if review and changed:
            row["Row_Status"] = "REVIEW"
            existing_error = clean_text(row.get("Error_Detail"))
            review_text = "Review assumptions before production load."
            if review_text not in existing_error:
                row["Error_Detail"] = f"{existing_error}; {review_text}".strip("; ")


def collect_files(inputs: Sequence[Path], explicit_files: Sequence[Path]) -> List[Path]:
    files: List[Path] = []
    for file_path in explicit_files:
        if file_path.exists() and file_path.is_file():
            files.append(file_path)
    for input_path in inputs:
        if input_path.is_file():
            files.append(input_path)
        elif input_path.exists():
            for suffix in ("*.xlsx", "*.xlsm", "*.pdf", "*.csv"):
                files.extend(input_path.rglob(suffix))
    unique: Dict[Path, Path] = {file.resolve(): file for file in files}
    return sorted(unique.values(), key=lambda path: str(path).lower())


def cell(ws: Any, row: int, col: int) -> Any:
    return ws.cell(row=row, column=col).value


def find_section(ws: Any, starts_with: str) -> Tuple[Optional[int], Optional[int], Optional[int]]:
    for row_index in range(1, ws.max_row + 1):
        value = clean_text(cell(ws, row_index, 1))
        if value.startswith(starts_with):
            header_row = row_index + 1
            total_row = None
            for probe in range(header_row + 1, ws.max_row + 1):
                if clean_text(cell(ws, probe, 1)) == "Total:":
                    total_row = probe
                    break
                next_section = clean_text(cell(ws, probe, 1))
                if next_section in {
                    "Property Pulse",
                    "Lead Activity",
                    "Lead Conversions",
                    "Make Ready Status",
                }:
                    break
            return row_index, header_row, total_row
    return None, None, None


def section_rows(ws: Any, header_row: Optional[int], total_row: Optional[int]) -> Iterable[int]:
    if not header_row or not total_row:
        return []
    return range(header_row + 1, total_row)


def row_values(ws: Any, row: int, cols: Sequence[int]) -> List[Any]:
    return [cell(ws, row, col) for col in cols]


def parse_box_score(path: Path, state: RunState, fallback_month: str, force_month: bool) -> None:
    # These reports are small enough to load in memory. Normal mode is much
    # faster than read_only mode for the random cell access used by section maps.
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    rows_extracted = 0
    report_months: List[str] = []
    communities: List[str] = []

    try:
        for ws in workbook.worksheets:
            if ws.title == "Report Parameters":
                continue
            title = clean_text(cell(ws, 2, 1))
            if "Box Score" not in title and not clean_text(cell(ws, 3, 1)).lower().startswith("rise"):
                continue

            source_property = clean_text(cell(ws, 3, 1)) or ws.title
            detected_period = clean_text(cell(ws, 4, 1))
            report_month = fallback_month if force_month and fallback_month else parse_report_month(detected_period) or fallback_month
            community_id = map_community(source_property)
            if not community_id:
                state.add_exception("ERROR", path, ws.title, f"Unmapped Box Score community: {source_property}", "Add the alias to Community_Aliases before loading.")
                continue

            report_months.append(report_month)
            communities.append(community_id)

            _, availability_header, availability_total = find_section(ws, "Availability")
            if availability_header and availability_total:
                for row_index in section_rows(ws, availability_header, availability_total):
                    unit_type = clean_text(cell(ws, row_index, 1))
                    if not unit_type:
                        continue
                    state.detail_rows["boxscore_floor_plan_monthly"].append(
                        dict(
                            zip(
                                TABLE_SCHEMAS["boxscore_floor_plan_monthly"],
                                [
                                    report_month,
                                    community_id,
                                    unit_type,
                                    *row_values(ws, row_index, list(range(2, 21))),
                                    cell(ws, row_index, 23),
                                    cell(ws, row_index, 24),
                                    cell(ws, row_index, 25),
                                    cell(ws, row_index, 26),
                                    cell(ws, row_index, 27),
                                    cell(ws, row_index, 28),
                                    cell(ws, row_index, 29),
                                    cell(ws, row_index, 30),
                                    cell(ws, row_index, 31),
                                    cell(ws, row_index, 32),
                                    cell(ws, row_index, 33),
                                    cell(ws, row_index, 34),
                                    cell(ws, row_index, 35),
                                    cell(ws, row_index, 36),
                                    cell(ws, row_index, 37),
                                    path.name,
                                    ws.title,
                                ],
                            )
                        )
                    )
                    rows_extracted += 1

                total_units = to_intish(cell(ws, availability_total, 6)) or to_intish(cell(ws, availability_total, 8))
                occupied_units = to_intish(cell(ws, availability_total, 9))
                leased_units = to_intish(cell(ws, availability_total, 35))
                vacant_units = to_intish(cell(ws, availability_total, 10))
                total_scheduled = to_number(cell(ws, availability_total, 27))
                total_effective = to_number(cell(ws, availability_total, 29))
                concessions = None
                if total_scheduled is not None and total_effective is not None:
                    concessions = max(total_scheduled - total_effective, 0)

                monthly_updates = {
                    "Total_Units": total_units,
                    "Occupied_Units": occupied_units,
                    "Leased_Units": leased_units,
                    "Preleased_Units": max((leased_units or 0) - (occupied_units or 0), 0)
                    if leased_units is not None and occupied_units is not None
                    else None,
                    "Vacant_Units": vacant_units,
                    "Physical_Occupancy_%": cell(ws, availability_total, 17),
                    "Leased_%": cell(ws, availability_total, 18),
                    "Avg_Market_Rent": cell(ws, availability_total, 3),
                    "Actual_Rent": cell(ws, availability_total, 4),
                    "Concessions_$": concessions,
                    "Gross_Potential_Rent_$": cell(ws, availability_total, 26),
                }
                state.merge_monthly(
                    report_month,
                    community_id,
                    monthly_updates,
                    "Box Score",
                    path,
                    notes=[
                        "Box Score supplies operating metrics.",
                        "Actual_Rent uses Avg. Scheduled Rent unless Atlas defines actual rent differently.",
                    ],
                    review=False,
                )

            _, pulse_header, pulse_total = find_section(ws, "Property Pulse")
            if pulse_header and pulse_total:
                for row_index in section_rows(ws, pulse_header, pulse_total):
                    unit_type = clean_text(cell(ws, row_index, 1))
                    if not unit_type:
                        continue
                    state.detail_rows["boxscore_pulse_monthly"].append(
                        dict(
                            zip(
                                TABLE_SCHEMAS["boxscore_pulse_monthly"],
                                [report_month, community_id, unit_type, *row_values(ws, row_index, list(range(2, 14))), path.name, ws.title],
                            )
                        )
                    )
                    rows_extracted += 1
                state.merge_monthly(
                    report_month,
                    community_id,
                    {
                        "Move_Ins": cell(ws, pulse_total, 3),
                        "Move_Outs": cell(ws, pulse_total, 9),
                        "Renewals_Signed": cell(ws, pulse_total, 6),
                    },
                    "Box Score",
                    path,
                    notes=["Renewals_Signed uses Renewal Leases Approved from Property Pulse."],
                    review=False,
                )

            _, lead_header, lead_total = find_section(ws, "Lead Activity")
            if lead_header and lead_total:
                for row_index in section_rows(ws, lead_header, lead_total):
                    unit_type = clean_text(cell(ws, row_index, 1))
                    if not unit_type:
                        continue
                    state.detail_rows["boxscore_lead_monthly"].append(
                        dict(
                            zip(
                                TABLE_SCHEMAS["boxscore_lead_monthly"],
                                [report_month, community_id, unit_type, *row_values(ws, row_index, list(range(2, 16))), path.name, ws.title],
                            )
                        )
                    )
                    rows_extracted += 1
                tours = cell(ws, lead_total, 15)
                state.merge_monthly(
                    report_month,
                    community_id,
                    {
                        "Traffic": tours,
                        "Leads": cell(ws, lead_total, 2),
                        "Tours": tours,
                    },
                    "Box Score",
                    path,
                    notes=["Traffic defaults to First Visits/Tours from Lead Activity."],
                    review=False,
                )

            _, app_header, app_total = find_section(ws, "Lead Conversions")
            if app_header and app_total:
                for row_index in section_rows(ws, app_header, app_total):
                    unit_type = clean_text(cell(ws, row_index, 1))
                    if not unit_type:
                        continue
                    state.detail_rows["boxscore_application_monthly"].append(
                        dict(
                            zip(
                                TABLE_SCHEMAS["boxscore_application_monthly"],
                                [report_month, community_id, unit_type, *row_values(ws, row_index, list(range(2, 11))), path.name, ws.title],
                            )
                        )
                    )
                    rows_extracted += 1
                applications = (to_number(cell(ws, app_total, 2)) or 0) + (to_number(cell(ws, app_total, 3)) or 0)
                state.merge_monthly(
                    report_month,
                    community_id,
                    {
                        "Applications": applications,
                        "Approved_Applications": cell(ws, app_total, 6),
                    },
                    "Box Score",
                    path,
                    notes=["Applications = Completed + Partially Completed from Lead Conversions."],
                    review=False,
                )

            _, make_ready_header, make_ready_total = find_section(ws, "Make Ready Status")
            if make_ready_header and make_ready_total:
                for row_index in section_rows(ws, make_ready_header, make_ready_total):
                    status = clean_text(cell(ws, row_index, 1))
                    if not status:
                        continue
                    state.detail_rows["boxscore_make_ready_monthly"].append(
                        dict(
                            zip(
                                TABLE_SCHEMAS["boxscore_make_ready_monthly"],
                                [report_month, community_id, status, *row_values(ws, row_index, [2, 3, 4, 5]), path.name, ws.title],
                            )
                        )
                    )
                    rows_extracted += 1

            if not COMMUNITIES[community_id]["In_Current_Atlas_Master"]:
                state.add_exception(
                    "REVIEW",
                    path,
                    ws.title,
                    f"{source_property} maps to {community_id}, which is not in the current Atlas Community_Master.",
                    "Add this community to Community_Master or exclude it from production load.",
                )

    finally:
        workbook.close()

    state.add_file_log(
        path,
        "Box Score",
        detected_period=", ".join(sorted(set(report_months))),
        report_month=", ".join(sorted(set(report_months))),
        community_scope=", ".join(sorted(set(communities))),
        rows_extracted=rows_extracted,
        status="OK" if rows_extracted else "REVIEW",
        notes="Parsed Box Score workbook.",
    )


def subject_from_market_workbook(workbook: Any, path: Path) -> str:
    if "Property Summary" in workbook.sheetnames:
        ws = workbook["Property Summary"]
        value = clean_text(cell(ws, 4, 1))
        if value:
            return value
    stem = re.sub(r"^20\d{2}_\d{2}_\d{2}_\d{2}_\d{2}_", "", path.stem)
    stem = re.sub(r"_Market$", "", stem)
    return stem.replace("_", " ")


def workbook_updated_date(workbook: Any, path: Path) -> str:
    snapshot = parse_snapshot_date_from_filename(path)
    for sheet_name in ["Property Summary", "Visual Market Survey", "Floor Plan Data", "Unit Level Data"]:
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        for row in range(1, min(ws.max_row, 5) + 1):
            for col in range(1, min(ws.max_column, 5) + 1):
                text = clean_text(cell(ws, row, col))
                match = re.search(r"Updated on:\s*(\d{1,2}/\d{1,2}/\d{4})", text, flags=re.I)
                if match:
                    return iso_date(match.group(1))
    return snapshot


def parse_market_survey(path: Path, state: RunState, fallback_month: str, force_month: bool) -> None:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    rows_extracted = 0
    try:
        subject_property = subject_from_market_workbook(workbook, path)
        community_id = map_community(subject_property)
        snapshot_date = workbook_updated_date(workbook, path)
        report_month = fallback_month if force_month and fallback_month else parse_report_month(snapshot_date) or fallback_month
        if not community_id:
            state.add_exception("ERROR", path, "Market Survey", f"Unmapped market subject property: {subject_property}", "Add the alias to Community_Aliases before loading.")
            state.add_file_log(path, "Market Survey", report_month=report_month, snapshot_date=snapshot_date, status="ERROR", notes="Subject community was not mapped.")
            return

        if "Visual Market Survey" in workbook.sheetnames:
            ws = workbook["Visual Market Survey"]
            for row_index in range(5, ws.max_row + 1):
                source_property = clean_text(cell(ws, row_index, 1))
                distance = cell(ws, row_index, 2)
                if not is_market_survey_property_row(source_property, distance):
                    continue
                is_subject = "Yes" if normalize_name(source_property) == normalize_name(subject_property) else "No"
                row = dict(
                    zip(
                        TABLE_SCHEMAS["market_property_snapshot"],
                        [
                            snapshot_date,
                            report_month,
                            community_id,
                            source_property,
                            is_subject,
                            distance,
                            cell(ws, row_index, 3),
                            cell(ws, row_index, 4),
                            cell(ws, row_index, 5),
                            cell(ws, row_index, 6),
                            cell(ws, row_index, 7),
                            cell(ws, row_index, 8),
                            cell(ws, row_index, 9),
                            cell(ws, row_index, 10),
                            cell(ws, row_index, 11),
                            cell(ws, row_index, 12),
                            cell(ws, row_index, 13),
                            cell(ws, row_index, 14),
                            cell(ws, row_index, 15),
                            cell(ws, row_index, 16),
                            cell(ws, row_index, 17),
                            cell(ws, row_index, 18),
                            cell(ws, row_index, 19),
                            path.name,
                        ],
                    )
                )
                state.detail_rows["market_property_snapshot"].append(row)
                rows_extracted += 1
                if is_subject == "Yes":
                    total_units = to_intish(row["Total_Units"])
                    vacant_units = to_intish(row["Vacant_Units"])
                    occupied_units = total_units - vacant_units if total_units is not None and vacant_units is not None else None
                    state.merge_monthly(
                        report_month,
                        community_id,
                        {
                            "Total_Units": total_units,
                            "Occupied_Units": occupied_units,
                            "Vacant_Units": vacant_units,
                            "Leased_%": row["Leased_Pct"],
                            "Avg_Market_Rent": row["Asking_Rent"],
                            "Concessions_$": row["Concessions"],
                            "Applications": row["Applications_Last_30_Days"],
                        },
                        "Market Survey",
                        path,
                        notes=["Market Survey is a pricing/comp snapshot and only fills blanks in monthly upload."],
                        review=True,
                    )

        if "Floor Plan Data" in workbook.sheetnames:
            ws = workbook["Floor Plan Data"]
            for row_index in range(4, ws.max_row + 1):
                source_property = clean_text(cell(ws, row_index, 1))
                if not source_property:
                    continue
                state.detail_rows["market_floor_plan_snapshot"].append(
                    dict(
                        zip(
                            TABLE_SCHEMAS["market_floor_plan_snapshot"],
                            [
                                snapshot_date,
                                report_month,
                                community_id,
                                source_property,
                                clean_text(cell(ws, row_index, 2)),
                                cell(ws, row_index, 3),
                                cell(ws, row_index, 4),
                                cell(ws, row_index, 5),
                                cell(ws, row_index, 6),
                                cell(ws, row_index, 7),
                                cell(ws, row_index, 8),
                                cell(ws, row_index, 9),
                                cell(ws, row_index, 10),
                                cell(ws, row_index, 11),
                                cell(ws, row_index, 12),
                                cell(ws, row_index, 13),
                                cell(ws, row_index, 14),
                                cell(ws, row_index, 15),
                                cell(ws, row_index, 16),
                                path.name,
                            ],
                        )
                    )
                )
                rows_extracted += 1

        if "Unit Level Data" in workbook.sheetnames:
            ws = workbook["Unit Level Data"]
            for row_index in range(4, ws.max_row + 1):
                source_property = clean_text(cell(ws, row_index, 1))
                if not source_property:
                    continue
                state.detail_rows["market_unit_snapshot"].append(
                    dict(
                        zip(
                            TABLE_SCHEMAS["market_unit_snapshot"],
                            [
                                snapshot_date,
                                report_month,
                                community_id,
                                source_property,
                                cell(ws, row_index, 2),
                                cell(ws, row_index, 3),
                                cell(ws, row_index, 4),
                                cell(ws, row_index, 5),
                                cell(ws, row_index, 6),
                                cell(ws, row_index, 7),
                                cell(ws, row_index, 8),
                                iso_date(cell(ws, row_index, 9)),
                                iso_date(cell(ws, row_index, 10)),
                                cell(ws, row_index, 11),
                                cell(ws, row_index, 12),
                                cell(ws, row_index, 13),
                                cell(ws, row_index, 14),
                                cell(ws, row_index, 15),
                                cell(ws, row_index, 16),
                                cell(ws, row_index, 17),
                                cell(ws, row_index, 18),
                                path.name,
                            ],
                        )
                    )
                )
                rows_extracted += 1

        state.add_file_log(
            path,
            "Market Survey",
            detected_period=f"Updated on {snapshot_date}" if snapshot_date else "",
            report_month=report_month,
            snapshot_date=snapshot_date,
            community_scope=community_id,
            rows_extracted=rows_extracted,
            status="OK" if rows_extracted else "REVIEW",
            notes="Parsed market survey workbook.",
        )
    finally:
        workbook.close()


def is_month_sheet_name(name: str) -> bool:
    return bool(re.fullmatch(r"(January|February|March|April|May|June|July|August|September|October|November|December)\s+20\d{2}", name))


def parse_renewal_tracker(path: Path, state: RunState, fallback_month: str, force_month: bool, include_pii: bool) -> None:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    rows_extracted = 0
    report_months: List[str] = []
    community_scope = ""
    try:
        setup = workbook["SETUP"] if "SETUP" in workbook.sheetnames else None
        source_property = clean_text(cell(setup, 3, 2)) if setup else ""
        community_id = map_community(source_property or path.stem)
        if not community_id:
            state.add_exception(
                "ERROR",
                path,
                "SETUP",
                f"Unmapped Renewal Tracker community: {source_property or path.stem}",
                "Add the property alias to Community_Aliases before loading.",
            )
            state.add_file_log(path, "Renewal Tracker", status="ERROR", notes="Community was not mapped.")
            return

        community_scope = community_id
        for ws in workbook.worksheets:
            if ws.title in {"SETUP", "TEMPLATE"} or not is_month_sheet_name(ws.title):
                continue
            sheet_report_month = parse_report_month(ws.title)
            report_month = fallback_month if force_month and fallback_month else sheet_report_month or fallback_month
            report_months.append(report_month)
            renewal_rows = 0
            signed_rows = 0
            transfer_rows = 0
            ntv_rows = 0

            for row_index in range(4, ws.max_row + 1):
                name = clean_text(cell(ws, row_index, 1))
                unit = clean_text(cell(ws, row_index, 2))
                expiration = iso_date(cell(ws, row_index, 4))
                if not name and not unit and not expiration:
                    continue
                if name.startswith(("RETENTION RATE", "Legend", "Renewed", "Pending Signature", "NTV", "Transfer", "MTM")):
                    break
                if not expiration or not re.match(r"20\d{2}-\d{2}-\d{2}", expiration):
                    continue

                signed_date = iso_date(cell(ws, row_index, 17))
                transfer_date = iso_date(cell(ws, row_index, 18))
                ntv_date = iso_date(cell(ws, row_index, 19))
                renewal_rows += 1
                signed_rows += 1 if signed_date else 0
                transfer_rows += 1 if transfer_date else 0
                ntv_rows += 1 if ntv_date else 0

                state.detail_rows["renewal_tracker_monthly"].append(
                    dict(
                        zip(
                            TABLE_SCHEMAS["renewal_tracker_monthly"],
                            [
                                report_month,
                                community_id,
                                source_property,
                                ws.title,
                                name if include_pii else "",
                                unit,
                                cell(ws, row_index, 3),
                                expiration,
                                iso_date(cell(ws, row_index, 5)),
                                iso_date(cell(ws, row_index, 6)),
                                iso_date(cell(ws, row_index, 7)),
                                cell(ws, row_index, 8),
                                cell(ws, row_index, 9),
                                cell(ws, row_index, 10),
                                cell(ws, row_index, 11),
                                cell(ws, row_index, 12),
                                cell(ws, row_index, 13),
                                cell(ws, row_index, 14),
                                cell(ws, row_index, 15),
                                cell(ws, row_index, 16),
                                signed_date,
                                transfer_date,
                                ntv_date,
                                clean_text(cell(ws, row_index, 20)) if include_pii else "",
                                clean_text(cell(ws, row_index, 21)),
                                cell(ws, row_index, 22),
                                cell(ws, row_index, 23),
                                cell(ws, row_index, 24),
                                cell(ws, row_index, 25),
                                cell(ws, row_index, 26),
                                cell(ws, row_index, 27),
                                cell(ws, row_index, 28),
                                "Yes" if include_pii else "No",
                                path.name,
                                ws.title,
                            ],
                        )
                    )
                )
                rows_extracted += 1

            if renewal_rows:
                state.merge_monthly(
                    report_month,
                    community_id,
                    {
                        "Renewals_Sent": renewal_rows,
                        "Renewals_Signed": signed_rows,
                    },
                    "Renewal Tracker",
                    path,
                    notes=[
                        "Renewals_Sent uses count of valid renewal tracker rows.",
                        f"Transfer rows={transfer_rows}; NTV rows={ntv_rows} retained in renewal detail table.",
                    ],
                    review=False,
                )

        state.add_file_log(
            path,
            "Renewal Tracker",
            detected_period=", ".join(sorted(set(report_months))),
            report_month=", ".join(sorted(set(report_months))),
            community_scope=community_scope,
            rows_extracted=rows_extracted,
            status="OK" if rows_extracted else "REVIEW",
            notes="Parsed multi-month renewal tracker workbook.",
        )
    finally:
        workbook.close()


def parse_date_range_month(date_range: str) -> str:
    first = clean_text(date_range).split(" - ", 1)[0]
    return parse_report_month(first)


def parse_trending_occupancy_csv(path: Path, state: RunState, fallback_month: str, force_month: bool) -> None:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))

    if not rows:
        state.add_file_log(path, "Trending Occupancy", status="REVIEW", notes="CSV had no data rows.")
        return

    first_range = clean_text(rows[0].get("Date Range"))
    blocks: List[List[Dict[str, str]]] = []
    current: List[Dict[str, str]] = []
    for row in rows:
        date_range = clean_text(row.get("Date Range"))
        if date_range == first_range and current:
            blocks.append(current)
            current = []
        current.append(row)
    if current:
        blocks.append(current)

    if len(blocks) == len(TRENDING_OCCUPANCY_COMMUNITY_ORDER):
        block_to_community = {
            index + 1: community_id
            for index, community_id in enumerate(TRENDING_OCCUPANCY_COMMUNITY_ORDER)
        }
        mapping_method = "Default RISE community block order"
    elif len(blocks) == 1 and map_community(path.stem):
        block_to_community = {1: map_community(path.stem) or ""}
        mapping_method = "Filename community alias"
    else:
        block_to_community = {}
        mapping_method = "Unmapped"
        state.add_exception(
            "ERROR",
            path,
            "Trending Occupancy",
            f"Expected {len(TRENDING_OCCUPANCY_COMMUNITY_ORDER)} community blocks or a filename community alias; found {len(blocks)} block(s).",
            "Add a community column to the source or update TRENDING_OCCUPANCY_COMMUNITY_ORDER.",
        )

    rows_extracted = 0
    report_months: List[str] = []
    communities: List[str] = []

    for block_index, block in enumerate(blocks, start=1):
        community_id = block_to_community.get(block_index, "")
        if community_id:
            communities.append(community_id)
        for row in block:
            source_month = parse_date_range_month(row.get("Date Range", ""))
            report_month = fallback_month if force_month and fallback_month else source_month or fallback_month
            report_months.append(report_month)
            detail_row = {
                "Report_Month": report_month,
                "Community_ID": community_id,
                "Date_Range": row.get("Date Range", ""),
                "Beginning_Occupancy": to_number(row.get("Beginning Occupancy")),
                "Move_Ins": to_intish(row.get("Move-ins")),
                "Move_Outs": to_intish(row.get("Move-outs")),
                "Ending_Occupancy": to_number(row.get("Ending Occupancy")),
                "Block_Index": block_index,
                "Mapping_Method": mapping_method,
                "Source_File_Name": path.name,
            }
            state.detail_rows["trending_occupancy_monthly"].append(detail_row)
            rows_extracted += 1

            if community_id:
                state.merge_monthly(
                    report_month,
                    community_id,
                    {
                        "Physical_Occupancy_%": detail_row["Ending_Occupancy"],
                        "Move_Ins": detail_row["Move_Ins"],
                        "Move_Outs": detail_row["Move_Outs"],
                    },
                    "Trending Occupancy",
                    path,
                    notes=["Trending Occupancy supplies multi-month occupancy, move-ins, and move-outs when Box Score is not present."],
                    review=False,
                )

    state.add_file_log(
        path,
        "Trending Occupancy",
        detected_period=", ".join(sorted(set(report_months))),
        report_month=", ".join(sorted(set(report_months))),
        community_scope=", ".join(sorted(set(communities))),
        rows_extracted=rows_extracted,
        status="OK" if rows_extracted and block_to_community else "REVIEW",
        notes=f"Parsed {len(blocks)} community block(s) using {mapping_method}.",
    )


def money_values(text: str) -> List[Optional[float]]:
    return [to_number(item) for item in re.findall(r"\(?-?\$?\d[\d,]*\.\d{2}\)?", text)]


def parse_pdf_property_header(lines: List[str]) -> Tuple[str, str]:
    for idx, line in enumerate(lines[:10]):
        if "03 - RISE - Rent Roll" in line:
            prop = line.split("03 - RISE - Rent Roll", 1)[1].strip()
            if not prop and idx + 1 < len(lines):
                prop = lines[idx + 1].strip()
            period = ""
            for candidate in lines[idx + 1 : idx + 4]:
                if re.match(r"^[A-Z][a-z]{2}\s+\d{4}$", candidate.strip()):
                    period = candidate.strip()
                    break
            return prop, period
        if line.startswith("Future Resident Details Property:"):
            prop = line.split("Property:", 1)[1].strip()
            return prop, ""
    return "", ""


RENT_ROLL_STATUSES = [
    "Occupied No Notice",
    "Notice Unrented",
    "Notice Rented",
    "Vacant Unrented",
    "Vacant Rented",
]


def parse_rent_roll_line(
    line: str,
    report_month: str,
    community_id: str,
    source_property: str,
    source_file: Path,
    page_number: int,
    include_pii: bool,
) -> Optional[Dict[str, Any]]:
    match = re.match(r"^([A-Za-z0-9.-]+)\s+(.+?)\s+([\d,]+\.\d{2})(.*)$", line)
    if not match:
        return None

    bldg_unit, unit_type, sqft, rest = match.groups()
    status = ""
    status_index = -1
    for candidate in RENT_ROLL_STATUSES:
        idx = rest.find(candidate)
        if idx >= 0 and (status_index < 0 or idx < status_index):
            status = candidate
            status_index = idx
    if status_index < 0:
        return None

    tail = rest[status_index + len(status) :].strip()
    ready_status = ""
    if status.startswith("Vacant"):
        ready_match = re.match(r"(Ready|Not Ready|Not\s+Ready)\s+", tail)
        if ready_match:
            ready_status = ready_match.group(1).replace("  ", " ")
            tail = tail[ready_match.end() :].strip()

    dates = re.findall(r"\d{2}/\d{2}/\d{4}", tail)
    first_date_pos = min([tail.find(date) for date in dates], default=-1)
    resident_name = ""
    if first_date_pos > 0:
        resident_name = tail[:first_date_pos].strip()

    amounts = money_values(tail)
    market_rent = actual_charges = in_place = balance = deposit = None
    if len(amounts) >= 5:
        market_rent, actual_charges, in_place, balance, deposit = amounts[-5:]
    elif len(amounts) >= 4:
        market_rent, actual_charges, in_place, balance = amounts[-4:]

    move_in = dates[0] if len(dates) >= 1 else ""
    lease_start = dates[1] if len(dates) >= 2 else ""
    lease_end = dates[2] if len(dates) >= 3 else ""
    expected_move_out = dates[3] if len(dates) >= 4 else ""

    return {
        "Report_Month": report_month,
        "Community_ID": community_id,
        "Source_Property_Name": source_property,
        "Bldg_Unit": bldg_unit,
        "Unit_Type": unit_type,
        "SqFt": to_number(sqft),
        "Unit_Status": status,
        "Resident_Name": resident_name if include_pii else "",
        "Move_In": iso_date(move_in),
        "Lease_Start": iso_date(lease_start),
        "Lease_End": iso_date(lease_end),
        "Expected_Move_Out": iso_date(expected_move_out),
        "Market_Rent_Budgeted": market_rent,
        "Actual_Charges": actual_charges,
        "In_Place_Rent_Scheduled": in_place,
        "Balance": balance if include_pii else "",
        "Deposit_Held": deposit if include_pii else "",
        "Ready_Status": ready_status,
        "PII_Load_Flag": "Yes" if include_pii else "No",
        "Source_File_Name": source_file.name,
        "Source_Page": page_number,
    }


def parse_rent_roll_pdf(path: Path, state: RunState, fallback_month: str, force_month: bool, include_pii: bool) -> None:
    if PdfReader is None:
        state.add_file_log(path, "Rent Roll PDF", status="ERROR", notes="pypdf is not installed.")
        state.add_exception("ERROR", path, "PDF", "Cannot parse PDF because pypdf is not installed.", "Install pypdf or run with the Codex bundled Python runtime.")
        return

    reader = PdfReader(str(path))
    current_property = ""
    current_period = ""
    rows_extracted = 0
    mapped_pages = 0
    report_months: List[str] = []
    communities: List[str] = []

    for page_index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        lines = [clean_text(line) for line in text.splitlines() if clean_text(line)]
        header_property, header_period = parse_pdf_property_header(lines)
        if header_property:
            current_property = header_property
            if header_period:
                current_period = header_period

        if not current_property:
            continue
        community_id = map_community(current_property)
        if not community_id:
            continue
        mapped_pages += 1
        report_month = fallback_month if force_month and fallback_month else parse_report_month(current_period) or fallback_month
        report_months.append(report_month)
        communities.append(community_id)

        for line in lines:
            parsed = parse_rent_roll_line(
                line,
                report_month,
                community_id,
                current_property,
                path,
                page_index,
                include_pii,
            )
            if parsed:
                state.detail_rows["rent_roll_unit_monthly"].append(parsed)
                rows_extracted += 1

    status = "OK" if rows_extracted else "REVIEW"
    notes = "PDF unit rows parsed with heuristic extraction; review wrapped lease/resident rows."
    state.add_file_log(
        path,
        "Rent Roll PDF",
        detected_period=", ".join(sorted(set(report_months))),
        report_month=", ".join(sorted(set(report_months))),
        community_scope=", ".join(sorted(set(communities))),
        rows_extracted=rows_extracted,
        status=status,
        notes=notes,
    )
    if mapped_pages and rows_extracted:
        state.add_exception(
            "REVIEW",
            path,
            "Rent Roll PDF",
            "PDF rows were extracted heuristically; wrapped names and split leases can require review.",
            "Review rent_roll_unit_monthly.csv before production load.",
        )
    elif mapped_pages and not rows_extracted:
        state.add_exception(
            "ERROR",
            path,
            "Rent Roll PDF",
            "Mapped rent-roll pages were found, but no unit rows were parsed.",
            "Review PDF layout and parser rules.",
        )


def detect_excel_type(path: Path) -> str:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        sheet_names = set(workbook.sheetnames)
        if "SETUP" in sheet_names and any(is_month_sheet_name(name) for name in sheet_names):
            return "Renewal Tracker"
        if {"Property Summary", "Floor Plan Data", "Unit Level Data"}.issubset(sheet_names):
            return "Market Survey"
        for ws in workbook.worksheets[:3]:
            title = clean_text(cell(ws, 2, 1))
            if "RISE - Box Score" in title or "Box Score" in title:
                return "Box Score"
        if "Monthly_Upload" in sheet_names:
            return "Atlas Template"
        return "Unknown Excel"
    finally:
        workbook.close()


def detect_csv_type(path: Path) -> str:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        try:
            header = next(reader)
        except StopIteration:
            return "Unknown CSV"
    normalized = [clean_text(item).lower() for item in header]
    if normalized == ["date range", "beginning occupancy", "move-ins", "move-outs", "ending occupancy"]:
        return "Trending Occupancy"
    return "Unknown CSV"


def parse_file(path: Path, state: RunState, fallback_month: str, force_month: bool, include_pii: bool, skip_pdf: bool) -> None:
    suffix = path.suffix.lower()
    try:
        if suffix in {".xlsx", ".xlsm"}:
            source_type = detect_excel_type(path)
            if source_type == "Box Score":
                parse_box_score(path, state, fallback_month, force_month)
            elif source_type == "Market Survey":
                parse_market_survey(path, state, fallback_month, force_month)
            elif source_type == "Renewal Tracker":
                parse_renewal_tracker(path, state, fallback_month, force_month, include_pii)
            elif source_type == "Atlas Template":
                state.add_file_log(path, "Atlas Template", status="SKIPPED", notes="Atlas template is an output schema, not a source report.")
            else:
                state.add_file_log(path, source_type, status="SKIPPED", notes="Unsupported Excel report shape.")
                state.add_exception("REVIEW", path, "Workbook", "Unsupported Excel report shape.", "Add parser support or exclude this file.")
        elif suffix == ".csv":
            source_type = detect_csv_type(path)
            if source_type == "Trending Occupancy":
                parse_trending_occupancy_csv(path, state, fallback_month, force_month)
            else:
                state.add_file_log(path, source_type, status="SKIPPED", notes="Unsupported CSV report shape.")
                state.add_exception("REVIEW", path, "CSV", "Unsupported CSV report shape.", "Add parser support or exclude this file.")
        elif suffix == ".pdf":
            if skip_pdf:
                state.add_file_log(path, "Rent Roll PDF", status="SKIPPED", notes="PDF parsing skipped by --skip-pdf.")
            else:
                parse_rent_roll_pdf(path, state, fallback_month, force_month, include_pii)
        else:
            state.add_file_log(path, "Unsupported", status="SKIPPED", notes=f"Unsupported extension {path.suffix}.")
    except Exception as exc:  # Keep the batch running and surface the failure.
        state.add_file_log(path, "Unknown", status="ERROR", notes=str(exc))
        state.add_exception("ERROR", path, "File", f"Failed to parse file: {exc}", "Open the source file and validate format.")


def csv_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        # Keep percentages and currency values stable without noisy binary decimals.
        return f"{value:.12g}"
    if isinstance(value, (dt.date, dt.datetime)):
        return iso_date(value)
    return str(value)


def write_csv(path: Path, fields: Sequence[str], rows: Sequence[Dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fields), extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field_name: csv_value(row.get(field_name, "")) for field_name in fields})


def write_combined_workbook(output_dir: Path, table_rows: Dict[str, Sequence[Dict[str, Any]]]) -> Path:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="0B3048")
    subheader_fill = PatternFill("solid", fgColor="DDECF4")
    header_font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=10, color="17202A")
    note_font = Font(name="Aptos", size=10, color="466579")
    thin_side = Side(style="thin", color="D6E0EA")
    border = Border(bottom=thin_side)
    tab_colors = {
        "monthly_upload": "0B3048",
        "exceptions": "A43A3A",
        "source_file_log": "6A7F8F",
        "manifest": "6A7F8F",
        "audit_move_in_out": "6A7F8F",
    }
    sheet_names = {
        "audit_move_in_out": "Audit_MoveInOut",
    }

    for table_name, rows in table_rows.items():
        fields = TABLE_SCHEMAS[table_name]
        sheet = workbook.create_sheet(sheet_names.get(table_name, table_name[:31]))
        sheet.sheet_properties.tabColor = tab_colors.get(table_name, "5D9AB5")
        sheet.append(fields)
        for cell_obj in sheet[1]:
            cell_obj.fill = header_fill
            cell_obj.font = header_font
            cell_obj.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_obj.border = border

        for row in rows:
            sheet.append([row.get(field_name, "") for field_name in fields])

        for row_cells in sheet.iter_rows(min_row=2):
            for cell_obj in row_cells:
                cell_obj.font = body_font
                cell_obj.alignment = Alignment(vertical="top", wrap_text=False)
                cell_obj.border = border

        if table_name in {"manifest", "exceptions", "source_file_log"}:
            for cell_obj in sheet[1]:
                cell_obj.fill = subheader_fill
                cell_obj.font = Font(name="Aptos", size=10, bold=True, color="0B3048")
            for row_cells in sheet.iter_rows(min_row=2):
                for cell_obj in row_cells:
                    cell_obj.font = note_font

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for col_index, field_name in enumerate(fields, start=1):
            max_len = len(field_name)
            for row_index in range(2, min(sheet.max_row, 80) + 1):
                value = sheet.cell(row=row_index, column=col_index).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            sheet.column_dimensions[get_column_letter(col_index)].width = min(max(max_len + 2, 12), 34)

    path = output_dir / "atlas_combined_upload.xlsx"
    workbook.save(path)
    return path


def finalize_monthly_rows(state: RunState) -> List[Dict[str, Any]]:
    rows = []
    for key in sorted(state.monthly_rows):
        row = state.monthly_rows[key]
        total_units = to_number(row.get("Total_Units"))
        occupied_units = to_number(row.get("Occupied_Units"))
        leased_units = to_number(row.get("Leased_Units"))
        physical_occupancy = to_number(row.get("Physical_Occupancy_%"))
        if row.get("Occupied_Units") in (None, "") and total_units is not None and physical_occupancy is not None:
            row["Occupied_Units"] = int(round(total_units * physical_occupancy))
            occupied_units = to_number(row.get("Occupied_Units"))
        if row.get("Physical_Occupancy_%") in (None, "") and total_units and occupied_units is not None:
            row["Physical_Occupancy_%"] = occupied_units / total_units
        if row.get("Leased_%") in (None, "") and total_units and leased_units is not None:
            row["Leased_%"] = leased_units / total_units
        if row.get("Vacant_Units") in (None, "") and total_units is not None and occupied_units is not None:
            row["Vacant_Units"] = max(total_units - occupied_units, 0)
        missing_required = [
            field_name
            for field_name in ("Report_Month", "Community_ID", "Occupied_Units", "Leased_Units")
            if row.get(field_name) in (None, "")
        ]
        if missing_required:
            row["Row_Status"] = "REVIEW"
            detail = clean_text(row.get("Error_Detail"))
            message = "Missing required monthly field(s): " + ", ".join(missing_required)
            if message not in detail:
                row["Error_Detail"] = f"{detail}; {message}".strip("; ")
        rows.append(row)
    return rows


def append_run_consistency_warnings(state: RunState) -> None:
    source_types = {
        clean_text(item.get("Detected_Source_Type"))
        for item in state.source_file_log
        if clean_text(item.get("Detected_Source_Type"))
    }
    if not source_types:
        return

    present_tables = {
        table_name
        for table_name, rows in state.detail_rows.items()
        if rows
    }
    has_core_ops_data = bool(
        source_types.intersection({"Box Score", "Market Survey", "Trending Occupancy", "Rent Roll", "Renewal Tracker"})
    )
    if not has_core_ops_data:
        return

    warning_specs = [
        (
            "Renewal Tracker",
            "renewal_tracker_monthly",
            "Renewal tracker input was not included in this run, so renewal offer and signed-renewal detail is missing from the package.",
            "Add the renewal tracker workbook when a full monthly package should include renewal detail.",
        ),
        (
            "Rent Roll",
            "rent_roll_unit_monthly",
            "Rent roll input was not included in this run, so unit-level resident and charge detail is missing from the package.",
            "Add the rent roll PDF when the monthly package should include unit-level rent roll detail.",
        ),
    ]

    existing_warnings = {
        (
            clean_text(item.get("Source_File_Name")),
            clean_text(item.get("Source_Location")),
            clean_text(item.get("Issue")),
        )
        for item in state.exceptions
    }
    for source_type, table_name, issue, action in warning_specs:
        if source_type in source_types or table_name in present_tables:
            continue
        key = ("RUN_SUMMARY", "Package Coverage", issue)
        if key in existing_warnings:
            continue
        state.exceptions.append(
            {
                "Severity": "WARN",
                "Source_File_Name": "RUN_SUMMARY",
                "Source_Location": "Package Coverage",
                "Issue": issue,
                "Recommended_Action": action,
            }
        )


def append_agent_note(row: Dict[str, Any], note: str) -> None:
    if not note:
        return
    existing = [item.strip() for item in str(row.get("Agent_Notes", "")).split(" | ") if item.strip()]
    if note not in existing:
        existing.append(note)
    row["Agent_Notes"] = " | ".join(existing)


def build_move_in_out_audit_rows(state: RunState) -> List[Dict[str, Any]]:
    trend_rows = state.detail_rows.get("trending_occupancy_monthly", [])
    trend_by_key: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for row in trend_rows:
        report_month = clean_text(row.get("Report_Month"))
        community_id = clean_text(row.get("Community_ID"))
        if not report_month or not community_id:
            continue
        trend_by_key[(report_month, community_id)] = row

    audit_rows: List[Dict[str, Any]] = []
    all_keys = sorted(set(state.monthly_rows.keys()) | set(trend_by_key.keys()))
    for key in all_keys:
        report_month, community_id = key
        monthly_row = state.monthly_rows.get(key)
        trend_row = trend_by_key.get(key)
        uploaded_move_ins = to_intish(monthly_row.get("Move_Ins")) if monthly_row else None
        uploaded_move_outs = to_intish(monthly_row.get("Move_Outs")) if monthly_row else None
        source_move_ins = to_intish(trend_row.get("Move_Ins")) if trend_row else None
        source_move_outs = to_intish(trend_row.get("Move_Outs")) if trend_row else None
        final_move_ins = uploaded_move_ins
        final_move_outs = uploaded_move_outs
        audit_status = "PASS"
        correction_note = ""

        if trend_row is None:
            audit_status = "FAIL"
            correction_note = "No occupancy trend source row was available for this community-month."
        else:
            if monthly_row is None:
                monthly_row = community_base_row(report_month, community_id)
                state.monthly_rows[key] = monthly_row
                state.monthly_priority[key] = SOURCE_PRIORITY["Trending Occupancy"]
            if uploaded_move_ins != source_move_ins or uploaded_move_outs != source_move_outs:
                audit_status = "FAIL"
                final_move_ins = source_move_ins
                final_move_outs = source_move_outs
                monthly_row["Move_Ins"] = source_move_ins if source_move_ins is not None else ""
                monthly_row["Move_Outs"] = source_move_outs if source_move_outs is not None else ""
                correction_note = (
                    f"Corrected Move_Ins/Move_Outs from {uploaded_move_ins if uploaded_move_ins is not None else 'blank'}"
                    f"/{uploaded_move_outs if uploaded_move_outs is not None else 'blank'} to "
                    f"{source_move_ins if source_move_ins is not None else 'blank'}/{source_move_outs if source_move_outs is not None else 'blank'} "
                    "from Trending Occupancy."
                )
                append_agent_note(monthly_row, correction_note)
                monthly_row["Row_Status"] = "REVIEW"
                detail = clean_text(monthly_row.get("Error_Detail"))
                mismatch_message = "Move_Ins/Move_Outs were corrected to match Trending Occupancy."
                if mismatch_message not in detail:
                    monthly_row["Error_Detail"] = f"{detail}; {mismatch_message}".strip("; ")
            else:
                final_move_ins = source_move_ins
                final_move_outs = source_move_outs

        audit_rows.append(
            {
                "Report_Month": report_month,
                "Community_ID": community_id,
                "Source_File_Name": clean_text(trend_row.get("Source_File_Name")) if trend_row else "",
                "Source_Move_Ins": source_move_ins if source_move_ins is not None else "",
                "Source_Move_Outs": source_move_outs if source_move_outs is not None else "",
                "Uploaded_Move_Ins": uploaded_move_ins if uploaded_move_ins is not None else "",
                "Uploaded_Move_Outs": uploaded_move_outs if uploaded_move_outs is not None else "",
                "Final_Move_Ins": final_move_ins if final_move_ins is not None else "",
                "Final_Move_Outs": final_move_outs if final_move_outs is not None else "",
                "Audit_Status": audit_status,
                "Correction_Note": correction_note,
            }
        )
    return audit_rows


def write_outputs(output_dir: Path, state: RunState) -> Dict[str, int]:
    output_dir.mkdir(parents=True, exist_ok=True)
    row_counts: Dict[str, int] = {}
    append_run_consistency_warnings(state)
    state.detail_rows["audit_move_in_out"] = build_move_in_out_audit_rows(state)

    monthly_rows = finalize_monthly_rows(state)
    table_rows: Dict[str, Sequence[Dict[str, Any]]] = {"monthly_upload": monthly_rows}
    write_csv(output_dir / "monthly_upload.csv", TABLE_SCHEMAS["monthly_upload"], monthly_rows)
    row_counts["monthly_upload"] = len(monthly_rows)

    for table_name, rows in state.detail_rows.items():
        table_rows[table_name] = rows
        write_csv(output_dir / f"{table_name}.csv", TABLE_SCHEMAS[table_name], rows)
        row_counts[table_name] = len(rows)

    write_csv(output_dir / "source_file_log.csv", TABLE_SCHEMAS["source_file_log"], state.source_file_log)
    table_rows["source_file_log"] = state.source_file_log
    row_counts["source_file_log"] = len(state.source_file_log)
    write_csv(output_dir / "exceptions.csv", TABLE_SCHEMAS["exceptions"], state.exceptions)
    table_rows["exceptions"] = state.exceptions
    row_counts["exceptions"] = len(state.exceptions)

    manifest_rows = [
        {"Table": table, "CSV_File": f"{table}.csv", "Rows": count}
        for table, count in sorted(row_counts.items())
    ]
    write_csv(output_dir / "manifest.csv", TABLE_SCHEMAS["manifest"], manifest_rows)
    table_rows["manifest"] = manifest_rows
    row_counts["manifest"] = len(manifest_rows)
    write_combined_workbook(output_dir, table_rows)

    (output_dir / "run_summary.json").write_text(json.dumps(row_counts, indent=2), encoding="utf-8")
    return row_counts


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Generate ATLAS-ready CSVs from RISE source reports.")
    parser.add_argument("-i", "--input", action="append", type=Path, default=[], help="Folder or file to scan. Can be provided multiple times.")
    parser.add_argument("-f", "--file", action="append", type=Path, default=[], help="Specific report file. Can be provided multiple times.")
    parser.add_argument("-o", "--output", type=Path, default=Path("outputs/atlas_uploads/latest"), help="Output folder for generated CSVs.")
    parser.add_argument("--report-month", default="", help="Fallback report month, e.g. 2026-04. Source dates win unless --force-report-month is used.")
    parser.add_argument("--force-report-month", action="store_true", help="Use --report-month for all parsed files regardless of source dates.")
    parser.add_argument("--include-pii", action="store_true", help="Include rent-roll resident names, balances, and deposits. Default masks these fields.")
    parser.add_argument("--skip-pdf", action="store_true", help="Skip PDF rent-roll parsing.")
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    fallback_month = report_month_from_cli(args.report_month) if args.report_month else ""
    if args.force_report_month and not fallback_month:
        parser.error("--force-report-month requires --report-month.")

    files = collect_files(args.input, args.file)
    if not files:
        parser.error("No report files found. Provide --input or --file.")

    state = RunState()
    for file_path in files:
        parse_file(file_path, state, fallback_month, args.force_report_month, args.include_pii, args.skip_pdf)

    row_counts = write_outputs(args.output, state)
    print(f"Wrote ATLAS upload package to {args.output.resolve()}")
    for table_name, count in sorted(row_counts.items()):
        print(f"  {table_name}: {count} rows")
    if state.exceptions:
        print(f"\nReview exceptions.csv before loading: {len(state.exceptions)} issue(s).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
